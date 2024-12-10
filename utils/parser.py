import os
import json
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup

def create_excel_file(excel_file):
    """
    Create a new Excel file with predefined columns.
    """
    wb = Workbook()
    ws = wb.active
    columns = [
        'ID', 'Title', 'Address', 'User Type', 'Square', 'Rooms', 'Status', 'Latitude', 'Longitude',
        'Price per m²', 'Created At', 'Owner Title', 'Owner Label Title', 'Owner Label Name',
        'Year of Construction', 'Condition', 'Floor', 'Number of Photos', 'Category Alias', 'District', 'Country',
        'City', 'Price'
    ]
    ws.append(columns)
    wb.save(excel_file)


def ensure_valid_excel_file(excel_file):
    """
    Ensure that the provided Excel file is valid.
    If it's not valid or doesn't exist, a new one will be created.
    """
    if os.path.exists(excel_file):
        try:
            # Attempt to open the existing file
            load_workbook(excel_file).close()
        except Exception:
            # If an error occurs, remove it and create a new one
            os.remove(excel_file)
            create_excel_file(excel_file)
    else:
        # Create a new file if it doesn't exist
        create_excel_file(excel_file)


def extract_data_from_script(soup):
    """
    Extract JSON data from the JavaScript section of the webpage.
    Specifically looks for 'window.data' in the script.
    """
    script = soup.find('script', string=lambda t: t and 'window.data' in t)
    if script:
        script_content = script.string.strip()
        json_data_start = script_content.find('{')
        json_data_end = script_content.rfind('}') + 1
        try:
            return json.loads(script_content[json_data_start:json_data_end])
        except json.JSONDecodeError:
            return None
    return None


def parse_detailed_page(data_id, excel_file):
    """
    Parse the detailed page of a listing by its ID.
    Extract and append all relevant information to the provided Excel file.

    :param data_id: The listing ID to parse.
    :param excel_file: The path to the Excel file.
    :return: True if data was successfully extracted and saved, False otherwise.
    """
    import requests
    detailed_url = f'https://krisha.kz/a/show/{data_id}'
    response = requests.get(detailed_url)
    soup = BeautifulSoup(response.text, 'html.parser')
    data = extract_data_from_script(soup)

    if data:
        advert = data.get('advert', {})
        adverts_list = data.get('adverts', [])

        advert_id = advert.get('id', 'N/A')
        title = advert.get('title', 'N/A')
        address = advert.get('addressTitle', 'N/A')
        user_type = advert.get('userType', 'N/A')
        square = advert.get('square', 'N/A')
        rooms = advert.get('rooms', 'N/A')
        status = advert.get('status', 'N/A')
        lat = advert.get('map', {}).get('lat', 'N/A')
        lon = advert.get('map', {}).get('lon', 'N/A')
        price = advert.get('price', 'N/A')
        district = advert.get('address', {}).get('district', 'N/A')
        country = advert.get('address', {}).get('country', 'N/A')
        city = advert.get('address', {}).get('city', 'N/A')

        if adverts_list:
            price_m2 = adverts_list[0].get('priceM2', 'N/A')
            created_at = adverts_list[0].get('createdAt', 'N/A')
            owner_data = adverts_list[0].get('owner', {})
            owner_title = owner_data.get('title', 'N/A')
            owner_label_title = owner_data.get('label', {}).get('title', 'N/A')
            owner_label_name = owner_data.get('label', {}).get('name', 'N/A')
            num_photos = adverts_list[0].get('nbPhotos', 'N/A')
            category_alias = adverts_list[0].get('category', {}).get('label', 'N/A')
        else:
            price_m2 = 'N/A'
            created_at = 'N/A'
            owner_title = 'N/A'
            owner_label_title = 'N/A'
            owner_label_name = 'N/A'
            num_photos = 'N/A'
            category_alias = 'N/A'

        construction_year, condition, floor = 'N/A', 'N/A', 'N/A'

        # Extract additional info such as construction year, floor, and condition
        for item in soup.find_all('div', class_='offer__info-item'):
            data_name = item.get('data-name', '')
            if data_name == 'house.year':
                construction_year = item.find('div', class_='offer__advert-short-info').text.strip()
            elif data_name == 'flat.floor':
                floor = item.find('div', class_='offer__advert-short-info').text.strip()
            elif data_name == 'flat.renovation':
                condition = item.find('div', class_='offer__advert-short-info').text.strip()

        wb = load_workbook(excel_file)
        ws = wb.active
        row = [
            advert_id, title, address, user_type, square, rooms, status, lat, lon,
            price_m2, created_at, owner_title, owner_label_title, owner_label_name,
            construction_year, condition, floor, num_photos, category_alias, district, country, city, price
        ]
        ws.append(row)
        wb.save(excel_file)
        return True
    return False
